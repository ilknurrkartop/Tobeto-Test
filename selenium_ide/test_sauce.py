from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c

class Test_SauceClass:
    
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 

    def teardown_method(self): 
        self.driver.quit()
        
    def test_blank_areas(self):
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username is required"
        
    def test_blank_password(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Password is required"
        
    def getData():
        excel = openpyxl.load_workbook(c.OUT_OF_SERVICE_XLSX)
        sheet = excel["Sheet1"]
        rows = sheet.max_row
        data=[]
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data
        
        
        
    
    @pytest.mark.parametrize("username,password",getData())
    def test_out_of_service_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH
        
    def test_invalid_login(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("locked_out_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."
        
        
        
    def test_successful_login(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        headerLogo = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='header_container']/div[1]/div[2]/div")))
        assert headerLogo.text == "Swag Labs"
        listOfCourses = self.driver.find_elements(By.CLASS_NAME, "inventory_item_label")
        assert len(listOfCourses) == 6
        
    def test_menu(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        menuShape = self.driver.find_element(By.ID,"react-burger-menu-btn")
        menuShape.click()
        link_element = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"logout_sidebar_link")))
        link_element.click()
        headerLogo = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='root']/div/div[1]")))
        assert headerLogo.text == "Swag Labs"
        
    def test_add_to_cart(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='add-to-cart-sauce-labs-backpack']")))
        addToCart.click()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='remove-sauce-labs-backpack']")))
        assert remove.text == "Remove"
        shopping_cart_link = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='shopping_cart_container']/a")))
        shopping_cart_link.click()
        shopping_cart_area = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='header_container']/div[2]/span")))
        assert shopping_cart_area.text == "Your Cart"
        
    def test_checkout(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='add-to-cart-sauce-labs-backpack']")))
        addToCart.click()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='remove-sauce-labs-backpack']")))
        assert remove.text == "Remove"
        shopping_cart_link = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='shopping_cart_container']/a")))
        shopping_cart_link.click()
        checkoutButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"checkout")))
        checkoutButton.click()
        shoppingInfo = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='header_container']/div[2]/span")))
        assert shoppingInfo.text == "Checkout: Your Information"
        
    def test_info(self):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='add-to-cart-sauce-labs-backpack']")))
        addToCart.click()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='remove-sauce-labs-backpack']")))
        assert remove.text == "Remove"
        shopping_cart_link = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='shopping_cart_container']/a")))
        shopping_cart_link.click()
        checkoutButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"checkout")))
        checkoutButton.click()
        firstnameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"first-name")))
        firstnameInput.send_keys("bürde")
        lastnameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"last-name")))
        lastnameInput.send_keys("kesikbaş")
        continueButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"continue")))
        continueButton.click()
        infoerrorMessage = self.driver.find_element(By.XPATH, "//*[@id='checkout_info_container']/div/form/div[1]/div[4]/h3")
        assert infoerrorMessage.text == "Error: Postal Code is required"